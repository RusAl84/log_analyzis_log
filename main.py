from PyQt5 import uic, QtCore, QtGui, QtWidgets
from openpyxl import load_workbook
from pyqtgraph import PlotWidget, plot
import pyqtgraph as pg
import sys  # We need sys so that we can pass argv to QApplication
import os
import numpy as np
from statsmodels.tsa.arima.model import ARIMA
import datetime
from datetime import timedelta
from pandas import read_csv
import pandas as pd
import csv


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        # Load UI
        uic.loadUi('main.ui', self)
        self.filename = 'logs.xlsx'  # Log file name

        self.pushButton.clicked.connect(self.load_logs)
        self.tableWidget.setColumnCount(5)
        self.tableWidget.setHorizontalHeaderLabels(
            ["Event ID", "Event Name", "Start Time", "End Time", "Duration (min)"])

        # Setup graph
        self.graphWidget.setBackground('w')
        self.graphWidget.setLabel(
            'left', 'Duration (min)', color='red', size=30)
        self.graphWidget.setLabel('bottom', 'Event ID', color='red', size=30)
        self.graphWidget.showGrid(x=True, y=True)

    def load_logs(self):
        # Load logs from Excel file
        wb = load_workbook(self.filename)
        sheet = wb.active

        logs = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            logs.append(row)

        # Fill table
        self.tableWidget.setRowCount(len(logs))
        for i, log in enumerate(logs):
            for j, value in enumerate(log):
                self.tableWidget.setItem(
                    i, j, QtWidgets.QTableWidgetItem(str(value)))

        # Analyze data
        self.analyze_logs(logs)

    def analyze_logs(self, logs):
        # Convert data to DataFrame for analysis
        df = pd.DataFrame(logs, columns=[
                          "Event ID", "Event Name", "Start Time", "End Time", "Duration (min)"])

        # Convert time to datetime
        df["Start Time"] = pd.to_datetime(df["Start Time"])
        df["End Time"] = pd.to_datetime(df["End Time"])
        df["Duration (min)"] = (df["End Time"] -
                                df["Start Time"]).dt.total_seconds() / 60

        # Plot event durations
        self.plot_event_durations(df)

        # Forecasting with ARIMA
        self.forecast_with_arima(df)

    def plot_event_durations(self, df):
        # Plotting event duration
        xdata = df["Event ID"].values
        ydata = df["Duration (min)"].values
        self.graphWidget.clear()
        self.graphWidget.plot(xdata, ydata, pen='r', symbol='o')

    def forecast_with_arima(self, df):
        # Forecasting using ARIMA
        df.set_index("Start Time", inplace=True)
        df = df.resample('D').sum()  # Resample to daily frequency
        df['Duration (min)'] = df['Duration (min)'].fillna(
            0)  # Fill NaN values

        # Fit ARIMA model
        model = ARIMA(df['Duration (min)'], order=(
            1, 1, 1))  # Adjust order as needed
        model_fit = model.fit()

        # Forecasting
        forecast = model_fit.forecast(steps=7)  # Forecast for the next 7 days
        forecast_index = pd.date_range(
            start=df.index[-1] + pd.Timedelta(days=1), periods=7, freq='D')

        # Plot forecast
        self.graphWidget.plot(forecast_index, forecast,
                              pen='g', symbol='x', name='Forecast')


def main():
    app = QtWidgets.QApplication(sys.argv)
    main = MainWindow()
    main.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
