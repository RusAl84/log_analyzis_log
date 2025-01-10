from openpyxl import load_workbook
import numpy as np
from random import randint
import datetime
from datetime import timedelta

filename='pokupki.xlsx'
wb = load_workbook(filename)

koef=np.zeros(29)
for i in range(0,29):
    str=wb["Sheet1"].cell(row=2+i, column=6).value
    koef[i] = int(int(str)/randint(18, 22))
koef[1]=koef[1]/4;
koef[2]=koef[2]*3;
koef[3]=koef[3]/4;
koef[4]=koef[4]*3;
koef[6]=koef[6]/4;
sheet = wb['Sheet1']

k = 30;
for z in range(0,k):
    for i in range(0, 29):          #row
        for c in range(1, 6):
            str=wb["Sheet1"].cell(row=2+i, column=c).value
            sheet.cell(31+29*z+i, c, str)
            str = wb["Sheet1"].cell(row=2 + i, column=7).value

            date_time_obj = datetime.datetime.strptime('2019.11.01', '%Y.%m.%d')
            date_time_obj=date_time_obj + timedelta(days=z+1)
            sheet.cell(31 + 29 * z + i, 7, date_time_obj)

            str=wb["Sheet1"].cell(row=2 + 29 * z + i, column=6).value;
            str=int(str)-randint(0, int(koef[i]))-randint(0, 1)*randint(0, 10)
            if str<0:
                str=0
            sheet.cell(31 + 29 * z + i, 6, str)

wb.save(filename)
